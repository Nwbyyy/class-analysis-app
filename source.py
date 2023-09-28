# main.py>

import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook

#import matplotlib.pyplot as plt
#from sklearn.preprocessing import StandardScaler
#from imblearn.over_sampling import RandomOverSampler

#pip install xlrd

def createarrays(worksheet, headers, num_columns):
        column_arrays = [[] for _ in range(num_columns)]
    
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            for col_idx, value in enumerate(row):
                column_arrays[col_idx].append(value)

        return column_arrays

def read_file(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    values = []
    headers = ['Student Name', 'Year', 'Major', 'GPA']

    num_columns = len(headers)
    column_arrays = createarrays(ws, headers, num_columns)
    print(column_arrays)
    wb.close()

read_file()
